import os
import streamlit as st
import re

from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.document_loaders import DirectoryLoader
from langchain.embeddings.sentence_transformer import SentenceTransformerEmbeddings
from langchain.vectorstores import FAISS

from langchain import PromptTemplate
from langchain.embeddings import HuggingFaceEmbeddings
from langchain.llms import CTransformers
from langchain.chains import RetrievalQA
from langchain.llms import LlamaCpp
from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler
from langchain.callbacks.manager import CallbackManager
from langchain.memory import ConversationBufferMemory
import time

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime


# sidebar contents
with st.sidebar:
        st.title('DOC-QA DEMO ')
        st.markdown('''
        ## About        
        This app is an LLM-powered Doc-QA Demo built using:
        - [Streamlit](https://streamlit.io/)
        - [LangChain](https://python.langchain.com/)
        - [HuggingFace](https://huggingface.co/declare-lab/flan-alpaca-large)
        ''')
        st.write ('Made this app for testing Document Question Answering with Custom URL Data')


custom_prompt_template = """ Use the following pieces of information to answer the user's question.
If you don't know the answer, please just say that you don't know the answer, don't try to make up
an answer. Please don't show unhelpful answers. If the question contains wrong spelling of person name, correct wrong spelling and show it in answer.

Context : {context}
Question : {question}

The answer should consist of at least 1 sentence for short questions or 7 sentences for more detailed qeustions. Only returns the helpful and reasonable answer below and nothing else.
No need to return the question. I just want answer. Please don't show unhelpful answers.
Helpful answer:
"""
def set_custom_prompt(custom_prompt_template):
    prompt = PromptTemplate(template=custom_prompt_template, input_variables=['context',
                                                                              'question'])
    return prompt

@st.cache_resource 
def load_llm():
    n_gpu_layers = 32  # Change this value based on your model and your GPU VRAM pool.
    n_batch = 512  # Should be between 1 and n_ctx, consider the amount of VRAM in your GPU.
    callback_manager = CallbackManager([StreamingStdOutCallbackHandler()])
    llm = LlamaCpp(
        #model_path="llama-2-7b-chat.ggmlv3.q8_0.bin",
        model_path="/home/sira/sira_project/meta-Llama2/llama-2-7b-chat.ggmlv3.q8_0.bin",
        n_gpu_layers=n_gpu_layers,
        n_batch=n_batch,
        callback_manager=callback_manager,
        verbose=True,n_ctx = 4096, temperature = 0.1, max_tokens = 4096
    )
    return llm


def check_duplicate(source_list):
    res = []
    for i in source_list:
        if i not in res:
            res.append(i)
    return res

def convert_to_website_format(urls):
    convert_urls = []
    for url in urls:
        # Remove any '.html' at the end of the URL
        url = re.sub(r'\.html$', '', url)
        # Check if the URL starts with 'www.' or 'http://'
        if not re.match(r'(www\.|http://)', url):
            url = 'https://' + url
        if '/index' in url:
            url = url.split('/index')[0]
        match = re.match(r'^([^ ]+)', url)
        if match:
            url = match.group(1)
        convert_urls.append(url)
    return convert_urls

def regex_source(answer):
    pattern = r"'source': '(.*?)'"
    matchs = re.findall(pattern, str(answer))
    convert_urls = convert_to_website_format(matchs)
    res_urls = check_duplicate(source_list=convert_urls)
    res_urls = filter_similar_url(res_urls)
    return res_urls

def filter_similar_url(urls):
    urls_remove = ["www.omniscien.com/aboutus/company","www.omniscien.com/lsev6/asr/automatic-speech-recognition-overview", "www.omniscien.com/lsev6/features/asr/autonomous-speech-recognition-overview","www.omniscien.com/lsev6/asr"]
    # Remove the URL from the list
    filtered_urls = [url for url in urls if url not in  urls_remove]
    return filtered_urls

def filter_search(db_similarity):
    filter_list = []
    top_score = db_similarity[0][1]
    for index, score in enumerate(db_similarity) :
        if score[1] - top_score <= 0.05:
              filter_list.append(score)
    return filter_list  


@st.cache_resource 
def load_embeddings():
    embeddings = HuggingFaceEmbeddings(model_name = "thenlper/gte-base",
                                       model_kwargs = {'device': 'cpu'})
    return embeddings

def log_to_excel(file_path, demo_version, question, answer, url_lst, response_time):
    try:
        # Load existing workbook or create a new one
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Timestamp", "Demo Version", "Question", "Answer", "URL", "Response Time"])

        # Add new entry to the sheet
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        urls_formatted = '\n'.join(url_lst)

        urls_list = urls_formatted.split('\n')
        for row_idx, url in enumerate(urls_list):
            cell = sheet.cell(row=sheet.max_row + row_idx + 1, column=5)
            hyperlink = Hyperlink(url, display=url)
            cell.hyperlink = hyperlink
            cell.font = Font(underline="single", color="0563C1")

        sheet.append([timestamp, demo_version, question, answer, urls_formatted, response_time])

        workbook.save(file_path)
        print("Logged successfully!")

    except Exception as e:
        print(f"An error occurred: {e}")


def main():
    fdate = datetime.now().strftime('%Y-%m-%d')
    file_path = "docqa_v-2.0_logs_"+fdate+".xlsx"
    demo_version = "2.0"
    
    global index
    st.header("DOCUMENT QUESTION ANSWERING FOR OMNISCIEN TECHNOLOGIES")
    st.subheader("Retrival QA (Running with GPU, max_token = 4096, K=3) ")
    #DB_FAISS_PATH = "./vectorstores_clean_doc_gte-base_no_overlap/db_faiss"
    DB_FAISS_PATH = "/home/sira/sira_project/meta-Llama2/vectorstores_clean_doc_gte-base/db_faiss"
    embeddings = load_embeddings()
    db = FAISS.load_local(DB_FAISS_PATH, embeddings)
    llm = load_llm()
    qa_prompt = set_custom_prompt(custom_prompt_template)
    memory = ConversationBufferMemory(memory_key="chat_history", 
                                      return_messages=True, 
                                      input_key="query", 
                                      output_key="result")
    qa_chain = RetrievalQA.from_chain_type(
        llm = llm,
        chain_type = "stuff",
        retriever = db.as_retriever(search_kwargs = {'k':3}), 
        return_source_documents = True,
        memory = memory,
        chain_type_kwargs = {"prompt":qa_prompt}) 

    query = st.text_input("ASK ABOUT THE DOCS:")        
    if query:
        start = time.time()
        db_similarity = db.similarity_search_with_score(query, k=5)
        filter_list = filter_search(db_similarity)
        response = qa_chain({'query': query})
        st.write(response["result"])
        urls = regex_source(filter_list)
        for count, url in enumerate(urls):
             st.write(str(count+1)+":", url)
             st.write(filter_list[count][1])
        end = time.time()
        st.write("Respone time:",int(end-start),"sec")

        log_to_excel(file_path, demo_version, query, response["result"], urls, str(int(end-start))+" sec")


if __name__ == '__main__':
        main()


